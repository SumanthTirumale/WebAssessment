from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.exceptions import IllegalCharacterError
from pathlib import Path

from assessment.helpers import create_folder


font = Font(bold=True, name='Calibri')

cell_alignment = Alignment(
    horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0
)

cell_border = Border(
    left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'),
)

yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')


class Reports:
    def __init__(self, **kwargs):
        self.path = create_folder()
        self.database = kwargs['database']
        self.vulnerability_type = kwargs['vulnerability_type']
        self.config_dict = kwargs['config_dict']
        self.start_date_time = kwargs['start_date_time'].strftime("%Y-%m-%d %H-%M-%S")
        self.end_date_time = kwargs['end_date_time'].strftime("%Y-%m-%d %H-%M-%S")
        self.book = Workbook()

    def get_reports(self):

        report_name = \
            f"{self.config_dict['printer_name']}_" \
            f"{self.config_dict['release_version']}_Web Assessment Report_{self.end_date_time}.xlsx"

        header_values = [
            'DATETIME', 'HOST', 'URL', 'METHOD', 'STATUS CODE', 'ATTACK TYPE',
            'TAG', 'ACTUAL DATA', 'PAYLOAD', 'CURRENT DATA', 'RESULT'
        ]

        self.execution_details()

        for assessment in self.vulnerability_type:

            if assessment == 'xss':

                xss_reports = self.database.get_reports('xss')

                self.book.create_sheet("XSS")

                sheet = self.book["XSS"]

                if len(xss_reports) > 0:
                    sheet.append(header_values)
                    self.add_data(sheet, xss_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

            if assessment == 'bufferoverflow':

                bufferoverflow_reports = self.database.get_reports('bufferoverflow')

                self.book.create_sheet("BufferOverFlow")

                sheet = self.book["BufferOverFlow"]

                if len(bufferoverflow_reports) > 0:
                    sheet.append(header_values)
                    self.add_data(sheet, bufferoverflow_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

            if assessment == 'secureheadercheck':

                active_header_audit_reports = self.database.get_reports('secureheadercheck')

                self.book.create_sheet("Active Header Audit")

                sheet = self.book["Active Header Audit"]

                if active_header_audit_reports is not None:
                    headers = ['Date Time', 'Host', 'URL', 'Method', 'Status code', 'Response Header',
                               'Vulnerabilities']
                    sheet.append(headers)
                    self.add_data(sheet, active_header_audit_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

            if assessment == 'csrf':

                xss_reports = self.database.get_reports('csrf')

                self.book.create_sheet("CSRF")

                sheet = self.book["CSRF"]

                if len(xss_reports) > 0:
                    sheet.append(header_values)
                    self.add_data(sheet, xss_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

            if assessment == 'directorytraversal':

                xss_reports = self.database.get_reports('dir_traversal')

                self.book.create_sheet("DirectoryTraversal")

                sheet = self.book["DirectoryTraversal"]

                if len(xss_reports) > 0:
                    sheet.append(header_values)
                    self.add_data(sheet, xss_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

            if assessment == 'authorizationbypass':
                xss_reports = self.database.get_reports('authorizationbypass')

                self.book.create_sheet("AuthorizationBypass")

                sheet = self.book["AuthorizationBypass"]

                if len(xss_reports) > 0:
                    sheet.append(header_values)
                    self.add_data(sheet, xss_reports)

                else:
                    sheet.merge_cells('C13:F13')
                    cell = sheet['C13']
                    cell.value = "No Vulnerabilities Found"
                    cell.alignment = cell_alignment
                    cell.font = font
                    cell.fill = yellow_fill

        self.book.save(Path(self.path)/report_name)
        return str(Path(self.path)/report_name)

    def execution_details(self):
        """Function to add execution details """

        sheet = self.book.active

        sheet.title = 'Execution Details'

        sheet['D13'] = "Printer Name"
        sheet['E13'] = self.config_dict['printer_name']

        sheet['D14'] = "Release Version"
        sheet['E14'] = self.config_dict['release_version']

        sheet['D15'] = "Protocol Type"
        sheet['E15'] = self.config_dict['protocol_type']

        sheet['D16'] = "Printer Type"
        sheet['E16'] = self.config_dict['printer_type']

        sheet['D17'] = "IP Address"
        sheet['E17'] = self.config_dict['ip_address']

        vulnerability_type = "\n".join(self.config_dict['security_risks'])
        sheet['D18'] = "Vulnerability Type"
        sheet['E18'] = vulnerability_type

        if type(self.config_dict['assessment_type']) is str:
            assessment_type = [self.config_dict['assessment_type']]
        else:
            assessment_type = self.config_dict['assessment_type']

        assessment_type = "\n".join(assessment_type)

        sheet['D19'] = "Assessment Type"
        sheet['E19'] = assessment_type

        if 'url_exclusion' in self.config_dict.keys():
            if type(self.config_dict['url_exclusion']) is str:
                url_exclusion = [self.config_dict['url_exclusion']]
            else:
                url_exclusion = self.config_dict['url_exclusion']

            url_exclusion = "\n".join(url_exclusion)
        else:
            url_exclusion = "No URL's Excluded"

        sheet['D20'] = "Excluded URL's"
        sheet['E20'] = url_exclusion

        if 'header_exclusion' in self.config_dict.keys():
            if type(self.config_dict['header_exclusion']) is str:
                header_exclusion = [self.config_dict['header_exclusion']]
            else:
                header_exclusion = self.config_dict['header_exclusion']

            header_exclusion = "\n".join(header_exclusion)
        else:
            header_exclusion = "No URL's Excluded"

        sheet['D21'] = "Excluded Headers"
        sheet['E21'] = header_exclusion

        sheet['D22'] = "Start Date"
        sheet['E22'] = self.start_date_time

        sheet['D23'] = "End Date"
        sheet['E23'] = self.end_date_time

        max_length = 0
        for col in sheet["D13:D23"]:
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

                    cell.alignment = cell_alignment
                    cell.border = cell_border
                    cell.font = font
                    cell.fill = yellow_fill

                except:
                    pass

            sheet.column_dimensions[column].width = max_length

        max_length = 0
        for col in sheet["E13:E23"]:
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

                    cell.alignment = Alignment(
                        horizontal='center', vertical='bottom', text_rotation=0, wrap_text=True,
                        shrink_to_fit=False, indent=0
                    )

                    cell.border = cell_border
                    cell.font = font

                except:
                    pass

            sheet.column_dimensions[column].width = max_length

    @staticmethod
    def add_data(sheet, data):
        for row in data:
            try:
                sheet.append(row)

            except IllegalCharacterError:

                tmp_list = list()

                for r in row:
                    tmp_list.append(r)

                tmp_list.pop(9)
                tmp_list.insert(9, "Character Is Not Supported")

                row = tuple(tmp_list)

                sheet.append(row)

        for cell in sheet["1:1"]:
            cell.alignment = cell_alignment
            cell.border = cell_border
            cell.font = font
            cell.fill = yellow_fill

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.column is 'H':
                        cell.alignment = Alignment(
                            horizontal='left', vertical='bottom',
                            text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0
                        )

                    else:
                        cell.alignment = Alignment(
                            horizontal='center', vertical='bottom',
                            text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0
                        )

                    cell.border = cell_border

                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

                except:
                    pass

            if column is 'A':
                sheet.column_dimensions[column].width = 25.00

            elif column is 'B':
                sheet.column_dimensions[column].width = 13.57

            elif column is 'C':
                sheet.column_dimensions[column].width = 43.57

            elif column is 'D' or column is 'E' or column is 'F':
                sheet.column_dimensions[column].width = 8.57

            elif column is 'G':
                sheet.column_dimensions[column].width = 13.57

            elif column is 'H':
                sheet.column_dimensions[column].width = 92.14

            elif column is 'I':
                sheet.column_dimensions[column].width = 23.57

            elif column is 'J':
                sheet.column_dimensions[column].width = 92.14

            else:
                sheet.column_dimensions[column].width = 19.57

